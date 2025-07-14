# Running instructions:
# Install the library openpyxl: pip install openpyxl

# Run this from the command line like:
# python  path\to\excel_file_maker.py  path\to\input_folder  path\to\excel_file_name

# It looks at all the files in the input folder.
# It deletes the information in the old excel file to create a new one.

import os
import sys
import openpyxl


def excel_maker(input_folder, excel_file_name):

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'TEMP'
    workbook.save(excel_file_name)

    print(os.listdir(input_folder))

    text_files = [x for x in os.listdir(input_folder) if x[-4:] == '.txt']

    print(text_files)

    for sheetnum, name in enumerate(text_files, 1):
        lines = []
        # Open file
        with open(os.path.join(input_folder, name), 'r') as f:
            # Read content of file
            lines += f.read().split('\n')

        excel_df = [x.split(' ') for x in lines]

        workbook = openpyxl.load_workbook(excel_file_name)
        workbook.create_sheet(f'{name}_{sheetnum}')
        sheet = workbook[f'{name}_{sheetnum}']

        for row in excel_df:
            sheet.append(row)

        # for row in excel_df:
        workbook.save(excel_file_name)

    workbook.remove(workbook['Sheet'])
    workbook.save(excel_file_name)

for arg in sys.argv:
    print(arg)

if len(sys.argv) == 3:
    excel_maker(sys.argv[1], sys.argv[2])
elif len(sys.argv) == 1:
    i = input("You are running it from the IDE. Interesting.")
    if i == 'Y':
        excel_maker('./file_input', './file_input/output.xlsx')
else:
    print('You have to write, in this order\n'
          '* "python"\n'
          '* "excel_file_maker.py" (the name of the python script)'
          '* the folder which contains the input\n'
          '* the new excel file being created\n')


