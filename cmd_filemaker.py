# Running instructions:
# Install the library openpyxl: pip install openpyxl

# Run this from the command line like:
# python  path\to\cmd_filemaker.py  path\to\input_folder  path\to\excel_file_name

# It looks at all the files in the input folder.
# It deletes the information in the old excel file to create a new one.

import os
import sys
import openpyxl


def excel_maker(input_folder, excel_file_name):

    # creates Excel workbook with temporary sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'TEMP'
    workbook.save(excel_file_name)

    # finds all .txt files
    text_files = [x for x in os.listdir(input_folder) if x[-4:] == '.txt']
    print("Files being checked:", text_files)

    for sheetnum, name in enumerate(text_files, 1):
        lines = []
        # read content of specific file
        with open(os.path.join(input_folder, name), 'r') as f:
            lines += f.read().split('\n')

        # the rows of the file correspond to Excel rows
        # the space-separated words correspond to Excel cells
        excel_df = [x.split(' ') for x in lines]

        # each file is put on its own sheet
        workbook = openpyxl.load_workbook(excel_file_name)
        workbook.create_sheet(f'{name}_{sheetnum}')
        sheet = workbook[f'{name}_{sheetnum}']
        for row in excel_df:
            sheet.append(row)

        # the Excel is saved now
        workbook.save(excel_file_name)

    # deletes temporary sheet
    workbook.remove(workbook['Sheet'])
    workbook.save(excel_file_name)


# this starts the actual code
if len(sys.argv) == 3:  # runs from command line correctly
    excel_maker(sys.argv[1], sys.argv[2])
else:  # incorrect somehow
    print('This is supposed to be run from the command line.\n'
          'You have to write, in this order:\n'
          '* "python"\n'
          '* "cmd_filemaker.py" (the name of the python script)\n'
          '* the folder which contains the input\n'
          '* the new excel file being created\n\n'
          'Assuming the path to the folder is "C:\\inputfolder" '
          'and the path to the file you want to create is "C:\\excelfile", '
          'it should look like:\n\n'
          'python cmd_filemaker.py C:\\inputfolder C:\\excelfile')


